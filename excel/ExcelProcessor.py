import openpyxl
import pandas
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from excel.AddressEntry import AddressEntry


class ExcelProcessor:
    def __init__(self):
        self.address_entries = dict()
        self.processed_file_name = None
        self.processed_work_book = None
        self.processed_work_sheet = None

        # TODO remove protected mode

    def pre_process_file(self, exported_file_name):
        self.processed_file_name = exported_file_name.split('.')[0] + '_Processed.xlsx'

        # Handle .xls format, which is actually HTML out of co-star
        if exported_file_name.split('.')[1] == 'xls':
            data = pandas.read_html(exported_file_name)[0]
            for i in range(1, len(data)):
                address_entry = AddressEntry(data[0][i])
                address_entry.set_zip_code(data[1][i])
                address_entry.set_leasing_company_name(data[7][i])
                self.address_entries[address_entry.address] = address_entry
        elif exported_file_name.split('.')[1] == 'csv':
            pass
        else:
            work_book = openpyxl.load_workbook(exported_file_name)
            work_sheet = work_book.get_sheet_by_name(exported_file_name.split('.')[0])

            for i in range(2, work_sheet.max_row+1):
                row = str(i)
                address_entry = AddressEntry(work_sheet['A' + row])
                address_entry.set_zip_code(work_sheet['B' + row])
                address_entry.set_leasing_company_name(work_sheet['H' + row])
                self.address_entries[address_entry.address] = address_entry

            work_book.close()

    def generate_post_processed_file(self):
        work_book = openpyxl.Workbook()
        work_sheet = work_book.active

        # Determine amount of max square footage columns and contacts
        square_footage_column_count = 0
        max_number_of_contacts = 0
        for key, value in self.address_entries.items():
            if len(value.square_footage) > square_footage_column_count:
                square_footage_column_count = len(value.square_footage)

            if len(value.contacts) > max_number_of_contacts:
                max_number_of_contacts = len(value.contacts)

        '''
        Establish columns
        '''
        # Building
        work_sheet.cell(1, 1).value = 'Building Address'
        work_sheet.column_dimensions[get_column_letter(1)].width = 40

        # Zip
        work_sheet.cell(1, 2).value = 'Zip'
        work_sheet.column_dimensions[get_column_letter(2)].width = 10

        # Square footage
        square_footage_column_start_idx = 3
        for i in range(square_footage_column_start_idx, square_footage_column_start_idx + square_footage_column_count):
            work_sheet.cell(1, i).value = 'Sf ' + str(i-2)

        # Rent
        rent_column_idx = square_footage_column_start_idx + square_footage_column_count
        work_sheet.cell(1, rent_column_idx).value = 'Actual Rent'
        work_sheet.column_dimensions[get_column_letter(rent_column_idx)].width = 15

        # Leasing company name
        leasing_company_idx = rent_column_idx + 1
        work_sheet.cell(1, leasing_company_idx).value = 'Leasing Company Name'
        work_sheet.column_dimensions[get_column_letter(leasing_company_idx)].width = 40

        # Contacts
        contacts_column_start_idx = leasing_company_idx + 1
        entry_idx = contacts_column_start_idx
        for i in range(0, max_number_of_contacts):
            contact_number = str(i+1)
            work_sheet.cell(1, entry_idx).value = 'Contact ' + contact_number
            work_sheet.cell(1, entry_idx+1).value = 'Email ' + contact_number
            work_sheet.cell(1, entry_idx+2).value = 'Phone ' + contact_number

            work_sheet.column_dimensions[get_column_letter(entry_idx)].width = 25
            work_sheet.column_dimensions[get_column_letter(entry_idx+1)].width = 25
            work_sheet.column_dimensions[get_column_letter(entry_idx+2)].width = 15
            entry_idx += 3

        # Style first row
        bold_font = Font(bold=True)
        for cell in work_sheet['1:1']:
            cell.font = bold_font

        '''
        Insert address data
        '''
        row = 2
        for address, address_entry in self.address_entries.items():
            # Address and Zip Code
            work_sheet.cell(row, 1).value = address
            work_sheet.cell(row, 2).value = address_entry.zip_code

            # Square Footage
            square_footage_idx = square_footage_column_start_idx
            for square_footage in address_entry.square_footage:
                work_sheet.cell(row, square_footage_idx).value = square_footage
                square_footage_idx += 1

            # Rent
            work_sheet.cell(row, rent_column_idx).value = address_entry.actual_rent

            # Leasing Company Name
            work_sheet.cell(row, leasing_company_idx).value = address_entry.leasing_company_name

            # Contact information
            contact_entry_idx = contacts_column_start_idx
            for contact in address_entry.contacts:
                work_sheet.cell(row, contact_entry_idx).value = contact.name
                work_sheet.cell(row, contact_entry_idx+1).value = contact.email
                work_sheet.cell(row, contact_entry_idx+2).value = contact.phone
                contact_entry_idx += 3

            row += 1

        work_book.save(self.processed_file_name)


if __name__ == '__main__':
    excel_processor = ExcelProcessor()
    excel_processor.pre_process_file('C:\\JensenProperties\\export.xls')
    excel_processor.address_entries['875 N Michigan Ave'].add_square_footage(4000)
    excel_processor.address_entries['875 N Michigan Ave'].add_square_footage(460)
    excel_processor.address_entries['875 N Michigan Ave'].add_contact('George', 'gk@gmail.com', '8479034782')
    excel_processor.address_entries['321 N Clark St'].add_contact('sam', 'sa@gmail.com', '8479034782')
    excel_processor.address_entries['321 N Clark St'].add_contact('ter', 'ter@gmail.com', '523-902-4452')
    excel_processor.generate_post_processed_file()
